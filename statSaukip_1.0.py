'''statSaukip_1.0
Updates:
- range of rows is automatic
Problems:
- не достаточно доступа для забора файлов с сервера этой прогой; тогда сначала нужно перенести интересующий год к себе на комп
'''
from csv import reader
from openpyxl import load_workbook, Workbook
from os import getcwd, listdir, chdir, remove, makedirs
from os.path import exists, isdir

def csv_to_xlsx(fileName):
	with open(fileName, newline='') as csvfile:
		wb = Workbook()
		ws = wb.active
		ws.title = 'convert'
		data = []

		CSVreader = reader(csvfile, delimiter=';', quotechar='|')

		for row in CSVreader:
			datarow = []
			for cell in row:
				datarow.append(''.join(cell))
			data.append(datarow)

		for rowd in data:
			ws.append(rowd)

		wb.save(r'{}\{}.xlsx'.format(getcwd(), ws['F2'].value[3:5:]))
		wb.close
		print('Конвертация {} успешно завершена'.format(fileName))

def getInfo(fileName):
	wb = load_workbook(filename=fileName, data_only=True)
	ws = wb.active
	y = ws['F2'].value[6:10:]
	n = ws['A2'].value
	info = [y, n]
	return info

def getHead(fileName):
	wb = load_workbook(filename=fileName, data_only=True)
	ws = wb.active
	head = [cell.value for cell in ws['A1':'AB1'][0]] 
	return head

def getData(wb):
	ws = wb.active
	data = []

	for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
		data.append(row)

	return data

# создаём рабочую папку для результатов
workFolder = r'{}\result'.format(getcwd())

if not exists(workFolder):
	makedirs(workFolder)

# находим все исходные csv
csvRoot = r'{}\2019'.format(getcwd())#input year!!!!!!!!!!
months = listdir(csvRoot)
csvFiles = []

fileType = input('''1 - объединить статистику
2 - объединить расход
иное - выход
Что выбираем? ->''')

for month in months:
	if isdir(r'{}\{}'.format(csvRoot, month)) == True:
		if fileType == '1':
			csvFiles.append(r'{}\{}\Статистика.csv'.format(csvRoot, month))
		elif fileType == '2':
			csvFiles.append(r'{}\{}\Расход.csv'.format(csvRoot, month))
		else:
			sys.exit

# конвертируем csv в xlsx с сохранием в рабочую папку
chdir(workFolder)

for file in csvFiles:
	csv_to_xlsx(file)

# находим все полученные xlsx
xlsxFiles = listdir(workFolder)
print('Объединяем файлы: \n{}'.format(xlsxFiles))

# создаём итоговый файл xlsx
year = getInfo(xlsxFiles[0])[0]
stendNumber = getInfo(xlsxFiles[0])[1]
wbr = Workbook()
wsr = wbr.active
wsr.title = "stend_{}".format(stendNumber)

# заполняем итоговый файл xlsx
head = getHead(xlsxFiles[0])
wsr.append(head)

for fileName in xlsxFiles:
	wb = load_workbook(filename=fileName, data_only=True)
	data = getData(wb)
	for row in data:
		wsr.append(row)
	wb.close

# сохраняем и закрываем итоговый файл, удаляем промежуточные файлы (.xlsx), уведомляем об успешности действий
resultFileName = 'stat_{}_{}.xlsx'.format(stendNumber, year)
wbr.save(resultFileName)
wbr.close

for file in xlsxFiles:
	remove(file)
	
print('Формирование файла статистики успешно завершено: {}'.format(resultFileName))
input()