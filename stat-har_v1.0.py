''' Stat-har_v1.0.py
Программа, формирующая характеристику из статистики.
Возникла такая надобность при внезапном обесточивании сервера во время испытаний,
файл характеристики не сохранился.'''
from openpyxl import load_workbook, Workbook
from math import pi

origin_wb = load_workbook(filename='Статистика.xlsx', data_only=True)
origin_ws = origin_wb.active
origin_firstRow = 14
origin_lastRow = 93
origin_dataSet = origin_ws[f'A{origin_firstRow}':f'AE{origin_lastRow}']

harakter_dataSet = []

for row in origin_dataSet:
	harakter_row = [row[11].value*79*pi/60000, row[8].value, row[9].value, row[10].value, row[15].value, row[19].value, row[16].value, row[17].value, row[18].value]
	harakter_dataSet.append(harakter_row)

# создаём файл для записи характеристики
result_wb = Workbook()
result_ws = result_wb.active
result_ws.title = 'Data'

# заполняем его
head = ['uk2пр','Gвпр','πк','ηкад','πт','Gгпр','ηт.е.','uт1/с0','μfт']
result_ws.append(head)

for row in harakter_dataSet:
	result_ws.append(row)

# сохраняем его и информируем об успехе
result_wb.save('UnitedRows.xlsx')
print('Формирование характеристики завершено')